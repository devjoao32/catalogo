import os
import pytest
from catalog.technical_specs import clear_technical_specs_cache
from catalog.onedrive import (
    _encode_share_url,
    categorize_photos,
    list_local_products,
    categorize_local_photos,
    find_local_images_for_code,
    resolve_local_asset_path,
    resolve_local_products_root,
)


def test_encode_share_url():
    url = "https://1drv.ms/f/c/abc123?e=xyz"
    encoded = _encode_share_url(url)
    assert encoded.startswith("u!")
    # Nao deve conter '='.
    assert "=" not in encoded


def test_categorize_photos():
    items = [
        {"name": "PROD1_branco.jpg", "webUrl": "http://example.com/w1"},
        {"name": "PROD1_ambient.jpg", "webUrl": "http://example.com/a1"},
        {"name": "PROD1_medida.jpg", "webUrl": "http://example.com/m1"},
        {"name": "OTHER.jpg", "webUrl": "http://example.com/o"},
    ]
    cats = categorize_photos(items, code="PROD1")
    assert cats["white_background"] == "http://example.com/w1"
    assert cats["ambient"] == "http://example.com/a1"
    assert cats["measures"] == "http://example.com/m1"

    # Filtro por codigo.
    cats2 = categorize_photos(items, code="OTHER")
    assert cats2["white_background"] is None
    assert cats2["ambient"] is None
    assert cats2["measures"] is None


def test_match_and_find_images(monkeypatch):
    # Prepara respostas falsas do Graph para simular pastas aninhadas.
    calls = {"share": 0, "children": 0}

    def fake_share_info(url):
        calls["share"] += 1
        return {"parentReference": {"driveId": "drive123"}, "id": "root"}

    def fake_list_children(drive_id, item_id):
        calls["children"] += 1
        if item_id == "root":
            return [
                {"name": "cat", "folder": {}, "id": "f1"},
                {"name": "6649.jpg", "id": "i1", "@microsoft.graph.downloadUrl": "u1"},
                {"name": "6649-1.jpg", "id": "i2", "@microsoft.graph.downloadUrl": "u2"},
                {"name": "other.png", "id": "i3", "@microsoft.graph.downloadUrl": "u3"},
            ]
        elif item_id == "f1":
            return [
                {"name": "6649_2.webp", "id": "i4", "@microsoft.graph.downloadUrl": "u4"},
                {"name": "junk.txt", "id": "i5"},
            ]
        else:
            return []

    # Aplica patch diretamente nas funcoes importadas no modulo onedrive.
    monkeypatch.setattr('catalog.onedrive.get_share_info', fake_share_info)
    monkeypatch.setattr('catalog.onedrive.list_children', fake_list_children)

    from catalog.onedrive import find_images_for_code

    imgs = find_images_for_code('https://1drv.ms/fake', '6649')
    assert len(imgs) == 3
    assert imgs[0]['name'] == '6649.jpg' and imgs[0]['variant'] == 0
    assert imgs[1]['name'] == '6649-1.jpg' and imgs[1]['variant'] == 1
    assert imgs[2]['name'] == '6649_2.webp' and imgs[2]['variant'] == 2

    # Segunda chamada deve usar cache; contadores nao devem aumentar.
    before = calls.copy()
    imgs2 = find_images_for_code('https://1drv.ms/fake', '6649')
    assert imgs2 == imgs
    assert calls == before


def test_auth_endpoints(monkeypatch):
    from fastapi.testclient import TestClient

    class DummyApp:
        def get_authorization_request_url(self, *args, **kwargs):
            assert kwargs.get("state")
            return "https://login.microsoftonline.com/fake"

        def acquire_token_by_authorization_code(self, code, scopes, redirect_uri):
            assert code == "abc123"
            return {"access_token": "token"}

    monkeypatch.setattr('catalog.auth._build_msal_app', lambda: DummyApp())

    # Importa app apos patch para registrar rotas normalmente.
    from app import app
    from catalog.auth import STATE_COOKIE_NAME

    client = TestClient(app)
    resp3 = client.get('/auth/login', params={'next': '/erp'}, follow_redirects=False)
    assert resp3.status_code in {302, 307}
    assert resp3.headers.get('location', '').startswith('https://login.microsoftonline.com')
    assert STATE_COOKIE_NAME in resp3.headers.get("set-cookie", "")

    state = client.cookies.get(STATE_COOKIE_NAME)
    assert state

    resp4 = client.get('/auth/callback')
    assert resp4.status_code == 400

    resp5 = client.get('/auth/callback', params={'code': 'abc123', 'state': 'invalid-state'})
    assert resp5.status_code == 400

    resp6 = client.get('/auth/callback', params={'code': 'abc123', 'state': state}, follow_redirects=False)
    assert resp6.status_code == 303
    assert resp6.headers.get('location') == '/erp'

    session = client.get('/auth/session')
    assert session.status_code == 200
    assert session.json()['authenticated'] is True
    assert session.json()['provider'] == 'azure'


def test_auth_callback_sanitizes_token_exchange_failure(monkeypatch):
    from fastapi.testclient import TestClient

    class DummyApp:
        def get_authorization_request_url(self, *args, **kwargs):
            return "https://login.microsoftonline.com/fake"

        def acquire_token_by_authorization_code(self, code, scopes, redirect_uri):
            return {"error": "invalid_grant", "error_description": "secret-provider-detail"}

    monkeypatch.setattr('catalog.auth._build_msal_app', lambda: DummyApp())

    from app import app
    from catalog.auth import STATE_COOKIE_NAME

    client = TestClient(app)
    login_response = client.get('/auth/login', follow_redirects=False)
    assert login_response.status_code in {302, 307}

    state = client.cookies.get(STATE_COOKIE_NAME)
    assert state

    callback_response = client.get('/auth/callback', params={'code': 'abc123', 'state': state})
    assert callback_response.status_code == 400
    assert callback_response.json() == {"detail": "OAuth token exchange failed"}


def test_local_products_and_local_photos(tmp_path):
    root = tmp_path / "01_PRODUTOS"
    product_dir = root / "ABAJUR" / "5989 - ABAJUR TESTE"
    product_dir.mkdir(parents=True)
    (product_dir / "foto_branco.jpg").write_bytes(b"white")
    (product_dir / "foto_ambient.jpg").write_bytes(b"ambient")
    (product_dir / "foto_medida.png").write_bytes(b"measures")

    products = list_local_products(str(root))
    assert len(products) == 1
    first = products[0]
    assert first["Codigo"] == "5989"
    assert first["Categoria"] == "ABAJUR"
    assert first["URLFoto"].startswith("/catalog/local/asset?path=")
    assert "foto_ambient.jpg" in first["URLFoto"]

    photos = categorize_local_photos("5989", str(root))
    assert photos["white_background"].startswith("/catalog/local/asset?path=")
    assert photos["ambient"].startswith("/catalog/local/asset?path=")
    assert photos["measures"].startswith("/catalog/local/asset?path=")


def test_local_products_enriches_with_cadastro_for_default_runtime(monkeypatch, tmp_path):
    one_drive_root = tmp_path / "OneDrive"
    photos_root = one_drive_root / "FOTOS_PRODUTOS"
    photos_root.mkdir(parents=True)
    (photos_root / "1580_1.jpg").write_bytes(b"img")

    monkeypatch.setenv("CATALOG_LOCAL_PRODUCTS_PATH", "")
    monkeypatch.setenv("OneDrive", str(one_drive_root))
    monkeypatch.setattr(
        "catalog.cadastro.load_cadastro_index",
        lambda path_override=None: {
            "1580": {
                "code": "1580",
                "name": "LAMP LED A60 BULBO 7W BIVOLT 3000K E27",
                "description": "Lâmpada LED Bulbo 7W A-60 Bivolt 3000K E27",
                "specs": "CÓDIGO: 1580 | POTÊNCIA: 7W",
                "category": "LUM SOBREPOR",
            }
        },
    )

    products = list_local_products()
    assert len(products) == 1
    item = products[0]
    assert item["Codigo"] == "1580"
    assert item["Nome"] == "LAMP LED A60 BULBO 7W BIVOLT 3000K E27"
    assert item["Descricao"] == "Lâmpada LED Bulbo 7W A-60 Bivolt 3000K E27"
    assert item["Especificacoes"] == "CÓDIGO: 1580 | POTÊNCIA: 7W"
    assert item["Categoria"] == "LUM SOBREPOR"


def test_local_products_fallbacks_to_stock_report(monkeypatch, tmp_path):
    from openpyxl import Workbook

    stock_report = tmp_path / "stock.xlsx"
    stock_photos = tmp_path / "stock-photos"
    stock_photos.mkdir(parents=True)
    (stock_photos / "1181 - FITA LED.png").write_bytes(b"img")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "POSICAO_ESTOQUE"
    worksheet.append(["FILIAL", "CODIGO", "DESCRICAO", "EMB"])
    worksheet.append([2, 2002, "LUMINARIA SPOT LS-001", "UND"])
    worksheet.append([2, 1181, "FITA LED 2835 3000K 12V 5M", "UND"])
    worksheet.append([2, 2002, "LUMINARIA DUPLICADA", "UND"])
    workbook.save(stock_report)
    workbook.close()

    monkeypatch.setenv("CATALOG_STOCK_REPORT_PATH", str(stock_report))
    monkeypatch.setenv("CATALOG_STOCK_PHOTOS_ROOT", str(stock_photos))
    monkeypatch.delenv("CATALOG_LOCAL_PRODUCTS_PATH", raising=False)
    monkeypatch.delenv("OneDrive", raising=False)
    monkeypatch.delenv("OneDriveCommercial", raising=False)
    monkeypatch.delenv("OneDriveConsumer", raising=False)
    monkeypatch.setenv("USERPROFILE", str(tmp_path / "user"))

    products = list_local_products()

    assert [item["Codigo"] for item in products] == ["1181", "2002", "2002"]
    assert products[0]["Nome"] == "FITA LED 2835 3000K 12V 5M"
    assert products[0]["Categoria"] == "FITA LED"
    assert products[1]["Nome"] == "LUMINARIA SPOT LS-001"
    assert products[1]["Categoria"] == "LUMINARIA"
    assert products[2]["Nome"] == "LUMINARIA DUPLICADA"
    assert products[0]["URLFoto"].startswith("/catalog/local/asset?path=")
    assert products[0]["FotoBranco"].startswith("/catalog/local/asset?path=")


def test_local_products_fill_specs_from_technical_specs_file(monkeypatch, tmp_path):
    root = tmp_path / "01_PRODUTOS"
    product_dir = root / "LAMPADAS" / "1580 - LAMP LED TESTE"
    product_dir.mkdir(parents=True)
    (product_dir / "1580_1.jpg").write_bytes(b"white")

    specs_file = tmp_path / "technical_specs_ascii.txt"
    specs_file.write_text(
        "CODIGO: 1580\\REFERENCIA: 1580\\POTENCIA(W): 7W\\BASE: E27\\GARANTIA: 2 anos",
        encoding="utf-8",
    )

    monkeypatch.setenv("CATALOG_TECHNICAL_SPECS_PATH", str(specs_file))
    clear_technical_specs_cache()

    products = list_local_products(str(root))
    assert len(products) == 1
    assert products[0]["Codigo"] == "1580"
    assert products[0]["Especificacoes"] == "CODIGO: 1580 | REFERENCIA: 1580 | POTENCIA(W): 7W | BASE: E27 | GARANTIA: 2 anos"
    return

    specs_file = tmp_path / "technical_specs.txt"
    specs_file.write_text(
        "CÓDIGO: 1580\\REFERÊNCIA: 1580\\POTÊNCIA(W): 7W\\BASE: E27\\GARANTIA: 2 anos",
        encoding="utf-8",
    )

    monkeypatch.setenv("CATALOG_TECHNICAL_SPECS_PATH", str(specs_file))
    clear_technical_specs_cache()

    products = list_local_products(str(root))
    assert len(products) == 1
    assert products[0]["Codigo"] == "1580"
    assert products[0]["Especificacoes"] == "CÓDIGO: 1580 | REFERÊNCIA: 1580 | POTÊNCIA(W): 7W | BASE: E27 | GARANTIA: 2 anos"
    monkeypatch.delenv("OneDriveConsumer", raising=False)
    monkeypatch.setenv("USERPROFILE", str(tmp_path / "user"))

    products = list_local_products()

    assert [item["Codigo"] for item in products] == ["1181", "2002", "2002"]
    assert products[0]["Nome"] == "FITA LED 2835 3000K 12V 5M"
    assert products[0]["Categoria"] == "FITA LED"
    assert products[1]["Nome"] == "LUMINARIA SPOT LS-001"
    assert products[1]["Categoria"] == "LUMINARIA"
    assert products[2]["Nome"] == "LUMINARIA DUPLICADA"
    assert products[0]["URLFoto"].startswith("/catalog/local/asset?path=")
    assert products[0]["FotoBranco"].startswith("/catalog/local/asset?path=")


def test_stock_report_gallery_uses_stock_photos(monkeypatch, tmp_path):
    from openpyxl import Workbook

    stock_report = tmp_path / "stock.xlsx"
    stock_photos = tmp_path / "stock-photos"
    stock_photos.mkdir(parents=True)
    (stock_photos / "1181 - FITA LED.png").write_bytes(b"img")
    (stock_photos / "1181-2.png").write_bytes(b"img2")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "POSICAO_ESTOQUE"
    worksheet.append(["FILIAL", "CODIGO", "DESCRICAO", "EMB"])
    worksheet.append([2, 1181, "FITA LED 2835 3000K 12V 5M", "UND"])
    workbook.save(stock_report)
    workbook.close()

    monkeypatch.setenv("CATALOG_STOCK_REPORT_PATH", str(stock_report))
    monkeypatch.setenv("CATALOG_STOCK_PHOTOS_ROOT", str(stock_photos))
    monkeypatch.delenv("CATALOG_LOCAL_PRODUCTS_PATH", raising=False)
    monkeypatch.delenv("OneDrive", raising=False)
    monkeypatch.delenv("OneDriveCommercial", raising=False)
    monkeypatch.delenv("OneDriveConsumer", raising=False)
    monkeypatch.setenv("USERPROFILE", str(tmp_path / "user"))

    images = find_local_images_for_code("1181")
    assert len(images) == 2
    assert all(image["url"].startswith("/catalog/local/asset?path=") for image in images)

    photos = categorize_local_photos("1181")
    assert photos["white_background"].startswith("/catalog/local/asset?path=")


def test_stock_report_matches_photo_by_description(monkeypatch, tmp_path):
    from openpyxl import Workbook

    stock_report = tmp_path / "stock.xlsx"
    stock_photos = tmp_path / "stock-photos"
    stock_photos.mkdir(parents=True)
    (stock_photos / "LUMINARIA SPOT LS-001 AMBIENTADA.psd").write_bytes(b"img")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "POSICAO_ESTOQUE"
    worksheet.append(["FILIAL", "CODIGO", "DESCRICAO", "EMB"])
    worksheet.append([2, 2002, "LUMINARIA SPOT LS-001", "UND"])
    workbook.save(stock_report)
    workbook.close()

    monkeypatch.setenv("CATALOG_STOCK_REPORT_PATH", str(stock_report))
    monkeypatch.setenv("CATALOG_STOCK_PHOTOS_ROOT", str(stock_photos))
    monkeypatch.delenv("CATALOG_LOCAL_PRODUCTS_PATH", raising=False)
    monkeypatch.delenv("OneDrive", raising=False)
    monkeypatch.delenv("OneDriveCommercial", raising=False)
    monkeypatch.delenv("OneDriveConsumer", raising=False)
    monkeypatch.setenv("USERPROFILE", str(tmp_path / "user"))

    products = list_local_products()
    assert len(products) == 1
    assert products[0]["Codigo"] == "2002"
    assert products[0]["URLFoto"].startswith("/catalog/local/asset?path=")

    images = find_local_images_for_code("2002")
    assert len(images) == 1
    assert images[0]["url"].startswith("/catalog/local/asset?path=")


def test_local_products_enriches_erp_only_items_with_stock_photos(monkeypatch, tmp_path):
    from openpyxl import Workbook

    one_drive_root = tmp_path / "OneDrive"
    photos_root = one_drive_root / "FOTOS_PRODUTOS"
    photos_root.mkdir(parents=True)
    (photos_root / "5989 - ABAJUR BASE.jpg").write_bytes(b"img")

    stock_report = tmp_path / "stock.xlsx"
    stock_photos = tmp_path / "stock-photos"
    stock_photos.mkdir(parents=True)
    (stock_photos / "1181 - FITA LED.png").write_bytes(b"img")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "POSICAO_ESTOQUE"
    worksheet.append(["FILIAL", "CODIGO", "DESCRICAO", "EMB"])
    worksheet.append([2, 1181, "FITA LED 2835 3000K 12V 5M", "UND"])
    workbook.save(stock_report)
    workbook.close()

    monkeypatch.setenv("CATALOG_LOCAL_PRODUCTS_PATH", "")
    monkeypatch.setenv("OneDrive", str(one_drive_root))
    monkeypatch.setenv("CATALOG_STOCK_REPORT_PATH", str(stock_report))
    monkeypatch.setenv("CATALOG_STOCK_PHOTOS_ROOT", str(stock_photos))
    monkeypatch.setattr(
        "catalog.erp_catalog.load_erp_index",
        lambda: {
            "1181": {
                "Codigo": "1181",
                "Nome": "FITA LED 2835 3000K 12V 5M",
                "Descricao": "FITA LED 2835 3000K 12V 5M",
                "Categoria": "FITA LED",
            }
        },
    )

    products = list_local_products()
    by_code = {item["Codigo"]: item for item in products}

    assert "1181" in by_code
    assert by_code["1181"]["URLFoto"].startswith("/catalog/local/asset?path=")
    assert by_code["1181"]["FotoBranco"].startswith("/catalog/local/asset?path=")


def test_local_products_accept_tiff(tmp_path):
    root = tmp_path / "Catalogo"
    product_dir = root / "ARANDELAS"
    product_dir.mkdir(parents=True)
    (product_dir / "6596 ARANDELA LED LUMIDOCK-001P.tif").write_bytes(b"tiff")
    (product_dir / "6596 - 2.tif").write_bytes(b"tiff2")

    products = list_local_products(str(root))
    by_code = {item["Codigo"]: item for item in products}

    assert "6596" in by_code
    assert by_code["6596"]["Categoria"] == "ARANDELA"
    assert by_code["6596"]["Nome"] == "ARANDELA LED LUMIDOCK-001P"
    assert by_code["6596"]["URLFoto"].startswith("/catalog/local/asset?path=")

    images = find_local_images_for_code("6596", str(root))
    assert len(images) == 2
    image_names = {image["name"] for image in images}
    assert "6596 ARANDELA LED LUMIDOCK-001P.tif" in image_names
    assert "6596 - 2.tif" in image_names


def test_resolve_local_asset_path(tmp_path):
    root = tmp_path / "01_PRODUTOS"
    folder = root / "A" / "1234 - PRODUTO"
    folder.mkdir(parents=True)
    file_path = folder / "img.jpg"
    file_path.write_bytes(b"img")

    resolved = resolve_local_asset_path("A/1234 - PRODUTO/img.jpg", str(root))
    assert resolved == os.path.abspath(str(file_path))

    blocked = resolve_local_asset_path("../outside.jpg", str(root))
    assert blocked is None


def test_local_index_ignores_empty_explicit_root(monkeypatch, tmp_path):
    empty_root = tmp_path / "empty"
    empty_root.mkdir(parents=True)
    real_root = tmp_path / "real" / "MARKETING" / "01_PRODUTOS" / "ABAJUR" / "5989 - TESTE"
    real_root.mkdir(parents=True)
    (real_root / "foto_branco.jpg").write_bytes(b"ok")

    monkeypatch.setenv("CATALOG_LOCAL_PRODUCTS_PATH", str(empty_root))
    monkeypatch.setenv("OneDrive", str(tmp_path / "real"))

    products = list_local_products()
    assert len(products) == 1
    assert products[0]["Codigo"] == "5989"


def test_resolve_local_root_prefers_ti1_catalogo(monkeypatch, tmp_path):
    one_drive_root = tmp_path / "OneDrive"
    premium_root = one_drive_root / "TI 1" / "catalogo"
    fotos_root = one_drive_root / "FOTOS_PRODUTOS"
    backup_root = one_drive_root / "MARKETING" / "01_PRODUTOS" / "BACKUP PRODUTOS"
    catalog_root = one_drive_root / "MARKETING" / "Catalogo"
    legacy_root = one_drive_root / "MARKETING" / "01_PRODUTOS"
    premium_root.mkdir(parents=True)
    fotos_root.mkdir(parents=True)
    backup_root.mkdir(parents=True)
    catalog_root.mkdir(parents=True)
    legacy_root.mkdir(parents=True, exist_ok=True)

    monkeypatch.setenv("CATALOG_LOCAL_PRODUCTS_PATH", "")
    monkeypatch.setenv("OneDrive", str(one_drive_root))

    resolved = resolve_local_products_root()
    assert resolved == os.path.abspath(str(premium_root))


def test_flat_backup_layout_groups_code_and_derives_category(tmp_path):
    root = tmp_path / "BACKUP PRODUTOS"
    root.mkdir(parents=True)
    (root / "1171 - FITA LED C FONTE 6500K NITROLUX (1).jpg").write_bytes(b"1")
    (root / "1171 - FITA LED C FONTE 6500K NITROLUX (2).jpg").write_bytes(b"2")
    (root / "1171 - FITA LED C FONTE 6500K NITROLUX (3).jpg").write_bytes(b"3")

    products = list_local_products(str(root))
    by_code = {item["Codigo"]: item for item in products}
    assert "1171" in by_code
    assert by_code["1171"]["Nome"] == "FITA LED C FONTE 6500K NITROLUX"
    assert by_code["1171"]["Categoria"] == "FITA LED"

    images = find_local_images_for_code("1171", str(root))
    names = [img["name"] for img in images]
    assert names == [
        "1171 - FITA LED C FONTE 6500K NITROLUX (1).jpg",
        "1171 - FITA LED C FONTE 6500K NITROLUX (2).jpg",
        "1171 - FITA LED C FONTE 6500K NITROLUX (3).jpg",
    ]


def test_flat_layout_supports_underscore_sequences(tmp_path):
    root = tmp_path / "FOTOS_PRODUTOS"
    root.mkdir(parents=True)
    (root / "1171_1.jpg").write_bytes(b"1")
    (root / "1171_2.jpg").write_bytes(b"2")
    (root / "1171_3.jpg").write_bytes(b"3")

    products = list_local_products(str(root))
    assert len(products) == 1
    assert products[0]["Codigo"] == "1171"
    assert products[0]["URLFoto"].endswith("1171_3.jpg")

    images = find_local_images_for_code("1171", str(root))
    names = [img["name"] for img in images]
    assert names == ["1171_1.jpg", "1171_2.jpg", "1171_3.jpg"]

    photos = categorize_local_photos("1171", str(root))
    assert photos["white_background"].endswith("1171_1.jpg")
    assert photos["measures"].endswith("1171_2.jpg")
    assert photos["ambient"].endswith("1171_3.jpg")


def test_underscore_sequences_keep_measures_without_ambient_fallback(tmp_path):
    root = tmp_path / "catalogo"
    root.mkdir(parents=True)
    (root / "1393_1.jpg").write_bytes(b"1")
    (root / "1393_2.jpg").write_bytes(b"2")

    photos = categorize_local_photos("1393", str(root))
    assert photos["white_background"].endswith("1393_1.jpg")
    assert photos["measures"].endswith("1393_2.jpg")
    assert photos["ambient"] is None


def test_category_variants_are_grouped_under_same_bucket(tmp_path):
    root = tmp_path / "BACKUP PRODUTOS"
    root.mkdir(parents=True)
    (root / "2001 - LUMINARIA SPOT LS-001 (1).jpg").write_bytes(b"1")
    (root / "2002 - LUMINARIA TARTARUGA LS-002 (1).jpg").write_bytes(b"1")
    (root / "2003 - LUMINARIA TRILHO LS-003 (1).jpg").write_bytes(b"1")

    products = list_local_products(str(root))
    by_code = {item["Codigo"]: item for item in products}
    assert by_code["2001"]["Categoria"] == "LUMINARIA"
    assert by_code["2002"]["Categoria"] == "LUMINARIA"
    assert by_code["2003"]["Categoria"] == "LUMINARIA"


def test_description_photo_is_prioritized_first(tmp_path):
    root = tmp_path / "Catalogo"
    product_dir = root / "ABAJUR" / "5989 - ABAJUR TOUCH"
    product_dir.mkdir(parents=True)
    (product_dir / "5989-2.jpg").write_bytes(b"2")
    (product_dir / "5989-3.jpg").write_bytes(b"3")
    (product_dir / "5989 - ABAJUR TOUCH (1).png").write_bytes(b"desc")

    images = find_local_images_for_code("5989", str(root))
    assert len(images) == 3
    assert images[0]["name"] == "5989 - ABAJUR TOUCH (1).png"


def test_variations_from_1_to_4_are_prioritized(tmp_path):
    root = tmp_path / "Catalogo"
    product_dir = root / "ABAJUR" / "5990 - ABAJUR CLASSIC"
    product_dir.mkdir(parents=True)
    (product_dir / "5990 - ABAJUR CLASSIC.png").write_bytes(b"main")
    (product_dir / "5990-1.jpg").write_bytes(b"v1")
    (product_dir / "5990 - 2.jpg").write_bytes(b"v2")
    (product_dir / "5990_4.jpg").write_bytes(b"v4")
    (product_dir / "5990-5.jpg").write_bytes(b"v5")

    images = find_local_images_for_code("5990", str(root))
    names = [img["name"] for img in images]

    assert names[0] == "5990 - ABAJUR CLASSIC.png"
    assert names[1:4] == ["5990-1.jpg", "5990 - 2.jpg", "5990_4.jpg"]
    assert names[4] == "5990-5.jpg"


def test_category_layout_integrates_all_products(tmp_path):
    root = tmp_path / "Catalogo"
    category = root / "ABAJUR"
    category.mkdir(parents=True)
    (category / "5989-ABAJUR TOUCH RECARREGAVEL ATS-001.jpg").write_bytes(b"a")
    (category / "5989-2.jpg").write_bytes(b"b")
    (category / "5990 - ABAJUR CLASSIC-001.png").write_bytes(b"c")
    (category / "5990 - 1.jpg").write_bytes(b"d")

    products = list_local_products(str(root))
    by_code = {item["Codigo"]: item for item in products}

    assert sorted(by_code.keys()) == ["5989", "5990"]
    assert by_code["5989"]["Categoria"] == "ABAJUR"
    assert by_code["5990"]["Categoria"] == "ABAJUR"
    assert by_code["5989"]["Nome"] == "ABAJUR TOUCH RECARREGAVEL ATS-001"
    assert by_code["5990"]["Nome"] == "ABAJUR CLASSIC-001"


def test_local_index_refreshes_after_folder_change(tmp_path):
    root = tmp_path / "Catalogo"
    category = root / "ABAJUR"
    category.mkdir(parents=True)
    (category / "5989-ABAJUR TOUCH.jpg").write_bytes(b"a")

    first = list_local_products(str(root))
    assert [item["Codigo"] for item in first] == ["5989"]

    (category / "5990-ABAJUR CLASSIC.jpg").write_bytes(b"b")
    second = list_local_products(str(root))
    codes = sorted(item["Codigo"] for item in second)
    assert codes == ["5989", "5990"]


def test_shortcut_entries_are_integrated_as_products(tmp_path):
    root = tmp_path / "Catalogo"
    category = root / "LUMI PERFIL"
    category.mkdir(parents=True)
    (category / "6379- LUMI PERFIL PE015B 2M 23.6x7mm - 17w BQ - 3000K.lnk").write_bytes(b"lnk")

    products = list_local_products(str(root))
    assert len(products) == 1
    item = products[0]
    assert item["Codigo"] == "6379"
    assert item["Categoria"] == "LUMINARIA"
    assert item["Nome"] == "LUMI PERFIL PE015B 2M 23.6x7mm - 17w BQ - 3000K"
    assert item["URLFoto"] == ""

    images = find_local_images_for_code("6379", str(root))
    assert images == []


def test_shortcut_suffix_is_removed_from_name(tmp_path):
    root = tmp_path / "Catalogo"
    category = root / "LUMI PERFIL"
    category.mkdir(parents=True)
    (category / "6380- LUMI PERFIL PE015B 3M 23.6x7mm - Atalho.lnk").write_bytes(b"lnk")

    products = list_local_products(str(root))
    assert len(products) == 1
    assert products[0]["Nome"] == "LUMI PERFIL PE015B 3M 23.6x7mm"


def test_shortcut_target_image_is_used(monkeypatch, tmp_path):
    drive_root = tmp_path / "drive"
    catalog_root = drive_root / "MARKETING" / "Catalogo"
    legacy_root = drive_root / "MARKETING" / "01_PRODUTOS"
    category = catalog_root / "ABAJUR"
    category.mkdir(parents=True)
    legacy_category = legacy_root / "ABAJUR" / "6001 - ABAJUR DROP-001B E27X1"
    legacy_category.mkdir(parents=True)

    shortcut_path = category / "6001 - ABAJUR DROP-001B E27X1.lnk"
    shortcut_path.write_bytes(b"lnk")
    target_image = legacy_category / "6001 - ABAJUR DROP-001B E27X1 (1).jpg"
    target_image.write_bytes(b"img")

    monkeypatch.setenv("CATALOG_LOCAL_PRODUCTS_PATH", "")
    monkeypatch.setenv("OneDrive", str(drive_root))

    monkeypatch.setattr(
        "catalog.onedrive._resolve_shortcut_targets",
        lambda scan_root: {os.path.abspath(str(shortcut_path)): os.path.abspath(str(target_image))},
    )

    products = list_local_products()
    by_code = {item["Codigo"]: item for item in products}
    assert "6001" in by_code
    assert by_code["6001"]["URLFoto"].startswith("/catalog/local/asset?path=")

    images = find_local_images_for_code("6001")
    assert len(images) == 1
    assert images[0]["name"] == "6001 - ABAJUR DROP-001B E27X1"


def test_local_products_merge_monthly_sales_from_stock_report(monkeypatch, tmp_path):
    from openpyxl import Workbook

    root = tmp_path / "Catalogo"
    category = root / "ABAJUR"
    category.mkdir(parents=True)
    (category / "1424 - BOCAL DE LOUCA M519 BIVOLT E27.jpg").write_bytes(b"img")

    report_path = tmp_path / "POSICAO_ESTOQUE.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "POSICAO_ESTOQUE"
    worksheet.append(
        [
            "FILIAL",
            "CODIGO",
            "DESCRICAO",
            "EMB",
            "UNI.",
            "CLASSE",
            "OBS2",
            "FISICO",
            "BLOQ.",
            "RESERV.",
            "DISP.",
            "ESTOQUE",
            "VL. ULT.ENT",
            "VENDA MES",
            "VENDA MES 1",
            "VENDA MES 2",
            "VENDA MES 3",
            "COD. BARRAS",
            "COD. FABRICA",
            "COD. MASTER",
            "COD. PRINCIPAL",
        ]
    )
    worksheet.append(
        [
            2,
            1424,
            "BOCAL DE LOUCA M519 BIVOLT E27",
            "UND",
            "UN",
            "C",
            "N",
            3669,
            0,
            0,
            3669,
            3669,
            0.351,
            600,
            50,
            0,
            0,
            7896688171513,
            "",
            1424,
            1424,
        ]
    )
    worksheet.append(
        [
            3,
            1424,
            "BOCAL DE LOUCA M519 BIVOLT E27",
            "UND",
            "UN",
            "C",
            "N",
            120,
            0,
            0,
            120,
            120,
            0.351,
            40,
            5,
            0,
            0,
            7896688171513,
            "",
            1424,
            1424,
        ]
    )
    workbook.save(report_path)
    workbook.close()

    monkeypatch.setenv("CATALOG_STOCK_REPORT_PATH", str(report_path))
    monkeypatch.setenv("CATALOG_STOCK_REPORT_AUTO_DISCOVERY", "0")
    monkeypatch.setenv("CATALOG_ERP_JSON_PATH", str(tmp_path / "missing_erp.json"))

    products = list_local_products(str(root))
    by_code = {item["Codigo"]: item for item in products}

    assert by_code["1424"]["VendaMes"] == 640
    assert by_code["1424"]["VendaMes1"] == 55
    assert by_code["1424"]["VendaMes2"] == 0
    assert by_code["1424"]["VendaMes3"] == 0

